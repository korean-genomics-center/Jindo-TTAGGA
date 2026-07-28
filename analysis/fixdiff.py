import openpyxl, itertools, statistics as st
F='Jindo_Supplementary_Data_v3.xlsx'
wb=openpyxl.load_workbook(F)
G=wb['S-Data 5 chrY genotypes']; R=[[c.value for c in r] for r in G.iter_rows()]
hdr=[str(x) if x is not None else '' for x in R[1]]
samples=hdr[5:25]
L=wb['S-Data 7 chrY lineages']; RL=[[c.value for c in r] for r in L.iter_rows()]
lin={str(r[0]):str(r[1]) for r in RL[2:] if r[0]}
A=[s for s in samples if lin.get(s)=='A']
B=[s for s in samples if lin.get(s)=='B' and s!='J73478']
print(f"lineage A {len(A)} / lineage B(J73478 제외) {len(B)}")
idx={s:i for i,s in enumerate(samples)}

rows=[]
for r in R[2:]:
    g=[str(r[5+i]).strip() for i in range(20)]
    va={g[idx[s]] for s in A if g[idx[s]]!='.'}
    vb={g[idx[s]] for s in B if g[idx[s]]!='.'}
    if len(va)==1 and len(vb)==1 and va!=vb:
        rows.append([r[0], r[1], r[2], r[3], list(va)[0], list(vb)[0]])
print(f"fixed differences = {len(rows)}  (기존 2,525)")

S=wb['S-Data 8 chrY fixed diffs']
wb.remove(S)
S=wb.create_sheet('S-Data 8 chrY fixed diffs')
S['A1']='Supplementary Data 8. Fixed differences between the two paternal lineages, counted across the 19 individuals retained after exclusion of the data-quality outlier J73478'
for j,h in enumerate(['chrom','pos','ref','alt','allele_lineage_A','allele_lineage_B'],1):
    S.cell(2,j,h)
for i,row in enumerate(rows,3):
    for j,v in enumerate(row,1): S.cell(i,j,v)
order=[n for n in ['S-Data 1 RefAbsent genes','S-Data 2 chrY genes','S-Data 3 per-chr SNP',
 'S-Data 4 SyRI variants','S-Data 5 chrY genotypes','S-Data 6 chrY distances',
 'S-Data 7 chrY lineages','S-Data 8 chrY fixed diffs','S-Data 9 chrY read metrics',
 'S-Data 10 telomere ends'] if n in wb.sheetnames]
wb._sheets=[wb[n] for n in order]
wb.save(F)
print("\n=== 최종 ===")
w2=openpyxl.load_workbook(F)
for n in w2.sheetnames:
    print(f"  {n:30s} {w2[n].max_row:7d} 행")
