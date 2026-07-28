import openpyxl, shutil, sys
from openpyxl.utils import get_column_letter

SRC = sys.argv[1] if len(sys.argv)>1 else 'Jindo_Supplementary_Data_v2.xlsx'
DST = 'Jindo_Supplementary_Data_v3.xlsx'
shutil.copy(SRC, DST)
wb = openpyxl.load_workbook(DST)

# 워드로 옮길 시트 (SI에 이미 대응 표 존재)
DROP = ['S-Table phasing (switch)', 'S-Table 20-sample WGS']

# 새 이름 매핑
RENAME = {
 'S-Data 1 RefAbsent genes' : 'S-Data 1 RefAbsent genes',
 'S-Data 2 chrY genes'      : 'S-Data 2 chrY genes',
 'S-Data 3 per-chr SNP'     : 'S-Data 3 per-chr SNP',
 'S-Data 4 SyRI variants'   : 'S-Data 4 SyRI variants',
 'S-Data 5a chrY genotypes' : 'S-Data 5 chrY genotypes',
 'S-Data 5b chrY distances' : 'S-Data 6 chrY distances',
 'S-Data 5c chrY lineages'  : 'S-Data 7 chrY lineages',
 'S-Data 5d chrY fixed diffs':'S-Data 8 chrY fixed diffs',
 'S-Data 5e chrY read metrics':'S-Data 9 chrY read metrics',
 'S-Table telomere ends'    : 'S-Data 10 telomere ends',
}
# 제목행(1행) 문구도 갱신
TITLE = {
 'S-Data 5 chrY genotypes'  : 'Supplementary Data 5. Per-site chrY genotypes for the 20 males',
 'S-Data 6 chrY distances'  : 'Supplementary Data 6. Pairwise chrY SNP distance matrix among the 20 males',
 'S-Data 7 chrY lineages'   : 'Supplementary Data 7. Paternal lineage assignment and within- and between-lineage distances for each individual',
 'S-Data 8 chrY fixed diffs': 'Supplementary Data 8. Fixed differences between the two paternal lineages',
 'S-Data 9 chrY read metrics':'Supplementary Data 9. Per-individual chrY read counts and depth against each reference',
 'S-Data 10 telomere ends'  : 'Supplementary Data 10. Telomeric repeat arrays at chromosome ends',
}

for s in DROP:
    if s in wb.sheetnames: del wb[s]; print(f"삭제: {s}")

for old, new in RENAME.items():
    if old in wb.sheetnames and old != new:
        wb[old].title = new; print(f"{old}  ->  {new}")

for name, title in TITLE.items():
    if name in wb.sheetnames:
        wb[name]['A1'] = title

order = ['S-Data 1 RefAbsent genes','S-Data 2 chrY genes','S-Data 3 per-chr SNP',
 'S-Data 4 SyRI variants','S-Data 5 chrY genotypes','S-Data 6 chrY distances',
 'S-Data 7 chrY lineages','S-Data 8 chrY fixed diffs','S-Data 9 chrY read metrics',
 'S-Data 10 telomere ends']
wb._sheets = [wb[n] for n in order if n in wb.sheetnames]
wb.save(DST)

wb2 = openpyxl.load_workbook(DST)
print("\n=== 최종 구조 ===")
for n in wb2.sheetnames:
    ws = wb2[n]
    print(f"  {n:30s} {ws.max_row:7d} x {ws.max_column:3d}   {str(ws['A1'].value)[:60]}")
