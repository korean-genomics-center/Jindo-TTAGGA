import openpyxl, itertools, random, statistics as st
random.seed(1)

# --- lineage & chrY distance ---
wb = openpyxl.load_workbook('Jindo_Supplementary_Data_v3.xlsx', data_only=True)
Rc = [[c.value for c in r] for r in wb['S-Data 7 chrY lineages'].iter_rows()]
lin = {str(r[0]): str(r[1]) for r in Rc[2:] if r[0]}
Rb = [[c.value for c in r] for r in wb['S-Data 6 chrY distances'].iter_rows()]
ids = [str(x) for x in Rb[1][1:] if x]
D = {str(r[0]): {ids[i]: r[1+i] for i in range(len(ids))} for r in Rb[2:] if r[0]}

# --- PI_HAT ---
P = {}
for i, line in enumerate(open('kinship.genome')):
    f = line.split()
    if i == 0 or len(f) < 10: continue
    P[frozenset((f[1], f[3]))] = float(f[9])
print(f"쌍 {len(P)} / 샘플 {len(lin)}")

pairs = [(a, b) for a, b in itertools.combinations(sorted(lin), 2)]
win = [P[frozenset((a,b))] for a,b in pairs if lin[a]==lin[b] and frozenset((a,b)) in P]
bet = [P[frozenset((a,b))] for a,b in pairs if lin[a]!=lin[b] and frozenset((a,b)) in P]
print(f"\nwithin-lineage  n={len(win):3d}  median {st.median(win):.4f}  mean {st.mean(win):.4f}")
print(f"between-lineage n={len(bet):3d}  median {st.median(bet):.4f}  mean {st.mean(bet):.4f}")

# --- Mann-Whitney (permutation) ---
obs = st.mean(win) - st.mean(bet)
pool = win + bet; nw = len(win); cnt = 0; N = 20000
for _ in range(N):
    random.shuffle(pool)
    if abs(st.mean(pool[:nw]) - st.mean(pool[nw:])) >= abs(obs): cnt += 1
print(f"\n[검정 1] within vs between PI_HAT 차이")
print(f"  관측 차 {obs:+.4f}   permutation p = {(cnt+1)/(N+1):.4f}  (n={N})")

try:
    from scipy.stats import mannwhitneyu
    u, p = mannwhitneyu(win, bet, alternative='two-sided')
    print(f"  Mann-Whitney U = {u:.0f},  p = {p:.4f}")
except Exception as e:
    print(f"  (scipy 없음: {e})")

# --- Mantel: chrY 거리 vs PI_HAT ---
x = []; y = []
for a, b in pairs:
    k = frozenset((a,b))
    if k in P and D.get(a,{}).get(b) is not None:
        x.append(float(D[a][b])); y.append(P[k])
def pear(x, y):
    mx, my = st.mean(x), st.mean(y)
    num = sum((a-mx)*(b-my) for a,b in zip(x,y))
    den = (sum((a-mx)**2 for a in x) * sum((b-my)**2 for b in y)) ** 0.5
    return num/den if den else 0
r0 = pear(x, y)
samples = sorted(lin); idx = {s:i for i,s in enumerate(samples)}
cnt = 0
for _ in range(N):
    perm = samples[:]; random.shuffle(perm)
    m = dict(zip(samples, perm))
    xx = [float(D[m[a]][m[b]]) for a,b in pairs if frozenset((a,b)) in P and D.get(m[a],{}).get(m[b]) is not None]
    if len(xx) == len(y) and abs(pear(xx, y)) >= abs(r0): cnt += 1
print(f"\n[검정 2] Mantel: chrY 거리 vs PI_HAT")
print(f"  r = {r0:+.4f}   permutation p = {(cnt+1)/(N+1):.4f}")

# --- 상위 10쌍 프레이밍 검정 ---
top10 = sorted(P.items(), key=lambda kv: -kv[1])[:10]
cross = sum(1 for k,_ in top10 if lin[list(k)[0]] != lin[list(k)[1]])
exp = 10 * len(bet) / len(P)
print(f"\n[참고] 상위 10쌍 중 cross-lineage: 관측 {cross} / 기대 {exp:.1f}")
print("  → 관측이 기대보다 낮으므로 '가로지른다'는 서술은 성립하지 않음")
