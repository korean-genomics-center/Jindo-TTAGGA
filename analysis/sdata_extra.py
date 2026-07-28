import openpyxl, itertools, statistics as st
F='Jindo_Supplementary_Data_v3.xlsx'
wb=openpyxl.load_workbook(F)

# lineage
RL=[[c.value for c in r] for r in wb['S-Data 7 chrY lineages'].iter_rows()]
lin={str(r[0]):str(r[1]) for r in RL[2:] if r[0]}

def load(tag):
    S=[l.strip() for l in open(f'{tag}_samples.txt')]
    rows=[]
    for line in open(f'{tag}_gt.tsv'):
        f=line.rstrip('\n').split('\t')
        gt=[x.replace('|','/').split('/')[0] for x in f[4:4+len(S)]]
        gt=['.' if x in ('.','') else x for x in gt]
        rows.append((f[0],int(f[1]),f[2],f[3],gt))
    return S,rows

def dist(S,rows,keep):
    idx={s:i for i,s in enumerate(S)}
    D={a:{b:0 for b in keep} for a in keep}
    for _,_,_,_,g in rows:
        for a,b in itertools.combinations(keep,2):
            x,y=g[idx[a]],g[idx[b]]
            if x!='.' and y!='.' and x!=y:
                D[a][b]+=1; D[b][a]+=1
    return D

def summary(D,keep,label,excl=None):
    k=[s for s in keep if s!=excl]
    A=[s for s in k if lin.get(s)=='A']; B=[s for s in k if lin.get(s)=='B']
    w=[D[a][b] for a,b in itertools.combinations(A,2)]+[D[a][b] for a,b in itertools.combinations(B,2)]
    bt=[D[a][b] for a in A for b in B]
    print(f"  {label:28s} within median {st.median(w):8.1f} | between {st.median(bt):8.1f} | ratio {st.median(bt)/st.median(w):5.2f}")
    return D

print("=== ROS_Cfam_1.0 ===")
S,rows=load('ros')
print(f"  필터 전 {len(rows)} 사이트")
rows_f=[r for r in rows if r[4].count('.')<=2]
print(f"  결측<=2 필터 후 {len(rows_f)} 사이트   (본문 14,792)")
Dr=dist(S,rows_f,S)
summary(Dr,S,'20명 전체')
summary(Dr,S,'J73478 제외',excl='J73478')

print("\n=== 4-Mb truncation control ===")
S2,rows2=load('trunc4Mb')
print(f"  {len(rows2)} 사이트")
Dt=dist(S2,rows2,S2)
summary(Dt,S2,'20명 전체')
summary(Dt,S2,'J73478 제외',excl='J73478')

# --- 시트 작성 ---
def add_matrix(name,title,D,keep):
    if name in wb.sheetnames: wb.remove(wb[name])
    ws=wb.create_sheet(name); ws['A1']=title
    for j,s in enumerate(keep,2): ws.cell(2,j,s)
    for i,a in enumerate(keep,3):
        ws.cell(i,1,a)
        for j,b in enumerate(keep,2): ws.cell(i,j,D[a][b])

add_matrix('S-Data 11 chrY dist ROS',
 'Supplementary Data 11. Pairwise chrY SNP distance matrix among the 20 males, computed against the ROS_Cfam_1.0 chrY (14,792 sites after missingness filtering)',Dr,S)
add_matrix('S-Data 12 chrY dist 4Mb',
 'Supplementary Data 12. Pairwise chrY SNP distance matrix among the 20 males, computed against the proximal 4 Mb of the Jindo1-G-TTAGGA chrY (2,109 sites), the length-matched truncation control',Dt,S2)

# PI_HAT
if 'S-Data 13 autosomal PI_HAT' in wb.sheetnames: wb.remove(wb['S-Data 13 autosomal PI_HAT'])
ws=wb.create_sheet('S-Data 13 autosomal PI_HAT')
ws['A1']='Supplementary Data 13. Autosomal identity-by-descent (PLINK PI_HAT) for all 190 pairs of the 20 males, with the chrY lineage of each individual'
for j,h in enumerate(['IID1','lineage1','IID2','lineage2','Z0','Z1','Z2','PI_HAT','same_lineage'],1): ws.cell(2,j,h)
i=3
for n,line in enumerate(open('kinship.genome')):
    f=line.split()
    if n==0 or len(f)<10: continue
    a,b=f[1],f[3]
    for j,v in enumerate([a,lin.get(a,''),b,lin.get(b,''),float(f[6]),float(f[7]),float(f[8]),float(f[9]),
                          'Y' if lin.get(a)==lin.get(b) else 'N'],1): ws.cell(i,j,v)
    i+=1
print(f"\nS-Data 13: {i-3} 쌍")

wb.save(F)
print("\n=== 최종 시트 ===")
for n in openpyxl.load_workbook(F).sheetnames: print("  "+n)
