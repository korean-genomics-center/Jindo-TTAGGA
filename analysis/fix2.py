import openpyxl, itertools, random, statistics as st
random.seed(1)
F='Jindo_Supplementary_Data_v3.xlsx'
wb=openpyxl.load_workbook(F)
RL=[[c.value for c in r] for r in wb['S-Data 7 chrY lineages'].iter_rows()]
lin={str(r[0]):str(r[1]) for r in RL[2:] if r[0]}

# ---------- 2-A : S-Data 11 재생성 (결측 <=1) ----------
S=[l.strip() for l in open('ros_samples.txt')]
rows=[]
for line in open('ros_gt.tsv'):
    f=line.rstrip('\n').split('\t')
    g=[x.replace('|','/').split('/')[0] for x in f[4:4+len(S)]]
    rows.append(['.' if x in ('.','') else x for x in g])
sub=[g for g in rows if g.count('.')<=1]
print(f"[2-A] 결측<=1 사이트 {len(sub)}  (본문 14,792)")
idx={s:i for i,s in enumerate(S)}
D={a:{b:0 for b in S} for a in S}
for g in sub:
    for a,b in itertools.combinations(S,2):
        x,y=g[idx[a]],g[idx[b]]
        if x!='.' and y!='.' and x!=y: D[a][b]+=1; D[b][a]+=1
for excl in (None,'J73478'):
    k=[s for s in S if s!=excl]
    A=[s for s in k if lin.get(s)=='A']; B=[s for s in k if lin.get(s)=='B']
    w=[D[a][b] for a,b in itertools.combinations(A,2)]+[D[a][b] for a,b in itertools.combinations(B,2)]
    bt=[D[a][b] for a in A for b in B]
    tag='20명' if excl is None else 'J73478제외'
    print(f"      {tag:10s} within {st.median(w):7.1f}  between {st.median(bt):7.1f}  ratio {st.median(bt)/st.median(w):.2f}")
n='S-Data 11 chrY dist ROS'
if n in wb.sheetnames: wb.remove(wb[n])
ws=wb.create_sheet(n)
ws['A1']='Supplementary Data 11. Pairwise chrY SNP distance matrix among the 20 males, computed against the truncated ROS_Cfam_1.0 chrY (14,792 sites after missingness filtering)'
for j,s in enumerate(S,2): ws.cell(2,j,s)
for i,a in enumerate(S,3):
    ws.cell(i,1,a)
    for j,b in enumerate(S,2): ws.cell(i,j,D[a][b])
order=[x for x in ['S-Data 1 RefAbsent genes','S-Data 2 chrY genes','S-Data 3 per-chr SNP','S-Data 4 SyRI variants',
 'S-Data 5 chrY genotypes','S-Data 6 chrY distances','S-Data 7 chrY lineages','S-Data 8 chrY fixed diffs',
 'S-Data 9 chrY read metrics','S-Data 10 telomere ends','S-Data 11 chrY dist ROS','S-Data 12 chrY dist 4Mb',
 'S-Data 13 autosomal PI_HAT'] if x in wb.sheetnames]
wb._sheets=[wb[x] for x in order]
wb.save(F)

# ---------- 2-B : 순열검정 P (중앙값 / 평균) ----------
P={}
for i,line in enumerate(open('kinship.genome')):
    f=line.split()
    if i==0 or len(f)<10: continue
    P[frozenset((f[1],f[3]))]=float(f[9])
samples=sorted(lin)
pairs=list(itertools.combinations(samples,2))
def stats(lab):
    w=[P[frozenset(p)] for p in pairs if lab[p[0]]==lab[p[1]]]
    b=[P[frozenset(p)] for p in pairs if lab[p[0]]!=lab[p[1]]]
    return st.median(w)-st.median(b), st.mean(w)-st.mean(b)
o_med,o_mean=stats(lin)
cm=cn=0; N=20000
for _ in range(N):
    perm=samples[:]; random.shuffle(perm)
    lab=dict(zip(samples,[lin[s] for s in perm]))
    m,a=stats(lab)
    if abs(m)>=abs(o_med): cm+=1
    if abs(a)>=abs(o_mean): cn+=1
print(f"\n[2-B] 관측 중앙값 차 {o_med:+.4f}  permutation P = {(cm+1)/(N+1):.3f}")
print(f"      관측 평균값 차 {o_mean:+.4f}  permutation P = {(cn+1)/(N+1):.3f}")
