#!/usr/bin/env python3
"""tblastn 집계 v2: query coverage를 HSP 병합으로 계산.
v1의 결함: best hit 하나만 봐서, exon별로 쪼개진 hit을 합산하지 못함.
v2: 같은 (query, subject) 쌍의 모든 HSP를 query 좌표에서 병합하여 총 coverage 산출.
"""
import glob, collections

OUT='${JINDO_ROOT}/Results/Gene_Annotation/RefAbsent_Functional/tblastn_genome'
TAGS=['ros','gsd','boxer','zoey']
ID_MIN=50.0; COV_MIN=50.0

queries={}
for line in open(f'{OUT}/query_3163.faa'):
    if line.startswith('>'):
        cur=line[1:].strip().split()[0]; queries[cur]=0
    else:
        queries[cur]+=len(line.strip())
print(f"query: {len(queries)}개\n")

def merge_intervals(iv):
    if not iv: return 0
    iv=sorted(iv)
    tot=0; s,e=iv[0]
    for a,b in iv[1:]:
        if a<=e+1: e=max(e,b)
        else: tot+=e-s+1; s,e=a,b
    tot+=e-s+1
    return tot

present={t:set() for t in TAGS}
bestcov=collections.defaultdict(dict)

for tag in TAGS:
    # (query, subject) -> [(qstart,qend), ...]  단, identity 기준 충족 HSP만
    hsp=collections.defaultdict(list)
    nhit=0
    for f in sorted(glob.glob(f'{OUT}/results_chunked/{tag}_q*.tsv')):
        for line in open(f):
            p=line.rstrip('\n').split('\t')
            if len(p)<10: continue
            q,s = p[0],p[1]
            pid=float(p[2]); qs=int(p[6]); qe=int(p[7])
            nhit+=1
            if pid>=ID_MIN:
                hsp[(q,s)].append((min(qs,qe),max(qs,qe)))
    # query별 최대 coverage (subject 하나 안에서 병합)
    qcov=collections.defaultdict(float)
    for (q,s),iv in hsp.items():
        covbp=merge_intervals(iv)
        cov=100.0*covbp/queries[q]
        if cov>qcov[q]: qcov[q]=cov
    for q,c in qcov.items():
        bestcov[q][tag]=c
        if c>=COV_MIN: present[tag].add(q)
    print(f"  {tag:6s}: {nhit:,} HSPs, present {len(present[tag]):,}")

seq_present=set()
for t in TAGS: seq_present |= present[t]
seq_absent=set(queries)-seq_present

print(f"\n{'='*66}")
print(f"■ tblastn 판정 v2 (identity>={ID_MIN}%, merged query coverage>={COV_MIN}%)")
print(f"{'='*66}")
print(f"  전체 reference-absent gene models : {len(queries):,}")
print(f"    ├─ sequence-present (게놈에 존재)  : {len(seq_present):,}  ({100*len(seq_present)/len(queries):.1f}%)")
print(f"    │    → annotation-absent")
print(f"    └─ sequence-absent  (게놈에도 없음): {len(seq_absent):,}  ({100*len(seq_absent)/len(queries):.1f}%)")

print(f"\n■ 게놈별 present")
for t in TAGS: print(f"    {t:6s}: {len(present[t]):,}")
all4=present['ros']&present['gsd']&present['boxer']&present['zoey']
print(f"\n  4개 게놈 전부에 존재: {len(all4):,}")

with open(f'{OUT}/sequence_absent_v2.txt','w') as o: o.write('\n'.join(sorted(seq_absent)))
with open(f'{OUT}/tblastn_classification_v2.tsv','w') as o:
    o.write("gene_id\tclass\t"+'\t'.join(f"{t}_merged_cov" for t in TAGS)+"\n")
    for q in sorted(queries):
        cls='sequence-absent' if q in seq_absent else 'annotation-absent'
        row=[q,cls]+[f"{bestcov[q].get(t,0):.1f}" for t in TAGS]
        o.write('\t'.join(row)+'\n')

print(f"\n■ 논문 대표 유전자 (v1 vs v2)")
for q in ['hap1_g2334.t1','hap2_g2255.t1']:
    cls='sequence-absent' if q in seq_absent else 'annotation-absent'
    cov=', '.join(f"{t}:{bestcov[q].get(t,0):.0f}%" for t in TAGS)
    print(f"    {q}: {cls}   [merged cov: {cov}]")

print(f"\n■ 발현 지지 교차")
try:
    import openpyxl
    wb=openpyxl.load_workbook('${JINDO_ROOT}/Results/Manuscript_Figures/Supplementary/data_xlsx/Jindo_Supplementary_Data.xlsx')
    ws=wb['S-Data 1 RefAbsent genes']; h=[c.value for c in ws[1]]
    gi=h.index('gene_id'); ei=h.index('expressed')
    expr={r[gi] for r in ws.iter_rows(min_row=2,values_only=True) if r[gi] and str(r[ei]).strip().upper()=='Y'}
    sa=seq_absent & expr
    print(f"    sequence-absent {len(seq_absent):,}개 중 발현: {len(sa):,}개 ({100*len(sa)/max(1,len(seq_absent)):.1f}%)")
except Exception as e:
    print(f"    (실패: {e})")
